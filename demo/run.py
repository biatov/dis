from openpyxl import Workbook
from openpyxl.styles import Font
import json
import subprocess


def read_xlsx():
    try:
        with open('data.json') as f:
            urls = json.load(f)
    except FileNotFoundError:
        urls = list()
    image = list(map(lambda el: el, urls))

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).value = 'Title'
    ws.cell(row=1, column=7).font = Font(bold=True)
    ws.cell(row=1, column=7).value = 'Image URL'

    for col, val in enumerate(image, start=2):
        ws.cell(row=col, column=2).value = val['title']
        ws.cell(row=col, column=7).value = val['image_url']
    try:
        wb.save("excel/store images.xlsx")
    except FileNotFoundError:
        pass

subprocess.call('scrapy crawl url -o data.json', shell=True)
read_xlsx()
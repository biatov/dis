from scrapy import Request
from scrapy.spiders import CrawlSpider
from ..items import DemoItem
from openpyxl import load_workbook


class AboutRestSpider(CrawlSpider):
    name = "url"

    allowed_domains = ["themastectomystore.com.au"]

    def read_xlsx(self):
        wb = load_workbook(filename='store images.xlsx')
        name_sheet = wb.get_sheet_names()[0]
        sheet = wb[name_sheet]
        for row in sheet.iter_rows():
            yield list(filter(None, [cell.value for cell in row]))

    def get_start_urls(self):
        return list(map(lambda i: ''.join(['http://www.themastectomystore.com.au/apps/search?q=', i[0].replace(' ', '+')]), list(self.read_xlsx())[1:]))

    def start_requests(self):
        for start_url in self.get_start_urls():
            yield Request(start_url, callback=self.parse, dont_filter=True)

    def parse(self, response):
        item = DemoItem()
        try:
            url_part = response.selector.xpath('.//ul[@id="wsite-search-product-results"]/li')[0].xpath('.//a/div/@style').extract_first().split("(")[1][:-1]
            item['image_url'] = ''.join([response.url.split('app')[0][:-1], url_part])
        except:
            item['image_url'] = ''
        item['title'] = response.url.split('=')[1].replace('+', ' ')
        yield item
